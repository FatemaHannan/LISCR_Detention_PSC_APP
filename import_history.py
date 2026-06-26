"""
Run this in your Codespace to bulk-import Consolidated Inspection History.
Usage: python3 import_history.py <path_to_xlsx>
"""
import sys, math
from openpyxl import load_workbook
import urllib.request, urllib.error, json

SUPABASE_URL = "https://gfrbwzyidlmanwmcmoua.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImdmcmJ3enlpZGxtYW53bWNtb3VhIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc4MDg3ODA3MSwiZXhwIjoyMDk2NDU0MDcxfQ.HPtVD0X6p9O-bfccAqxMSrFu4xSmJcMU3yR4h3zg8y0"
HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": "Bearer " + SUPABASE_KEY,
    "Content-Type": "application/json",
    "Prefer": "resolution=merge-duplicates,return=minimal"
}

def sb_request(method, path, data=None):
    url = SUPABASE_URL + "/rest/v1/" + path
    body = json.dumps(data).encode() if data else None
    req = urllib.request.Request(url, data=body, headers=HEADERS, method=method)
    try:
        with urllib.request.urlopen(req) as r:
            return r.status, r.read()
    except urllib.error.HTTPError as e:
        return e.code, e.read()

def clean_imo(v):
    if v is None: return ""
    num = round(float(v)) if isinstance(v, (int, float)) else None
    if num and 1000000 < num < 10000000:
        return str(num)
    digits = ''.join(c for c in str(v) if c.isdigit())
    return digits[-7:] if len(digits) > 7 else digits

def s(v): return "" if v is None else str(v).strip()
def n(v):
    try: return float(str(v).replace(',','')) if v is not None else 0
    except: return 0
def si(v):
    try: return int(float(str(v))) if v is not None else 0
    except: return 0
def d(v):
    if v is None: return None
    if hasattr(v, 'strftime'): return v.strftime('%Y-%m-%d')
    return str(v).strip()[:10] or None

file_path = sys.argv[1] if len(sys.argv) > 1 else "Consolidated_Inspection_History__1_.xlsx"
print(f"Loading {file_path}...")
wb = load_workbook(file_path, read_only=True, data_only=True)
ws = wb.active
headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]

rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    row = dict(zip(headers, r))
    vessel = s(row.get('Vessel'))
    imo_val = clean_imo(row.get('IMO#'))
    if not vessel or not imo_val: continue
    rows.append({
        'vessel': vessel, 'imo': imo_val,
        'inspection_date': d(row.get('Inspection Date')),
        'port': s(row.get('Port')), 'mou': s(row.get('MOU')),
        'flag_psc': s(row.get('Flag/PSC')), 'car_status': s(row.get('CAR Status')),
        'num_findings': si(row.get('#Findings')),
        'detainable_flag': s(row.get('Detainable Flag')),
        'finding_note': s(row.get('Finding Note')),
        'was_detained': s(row.get('Was Detained')),
        'inspection_type': s(row.get('Inspection Type')),
        'days_since_last': n(row.get('Days')),
        'last_onboard': s(row.get('Last Onboard')), 'auditor': s(row.get('Auditor')),
        'ism_client': s(row.get('ISM Client')), 'risk_level': s(row.get('Risk Level')),
        'target_vessel': s(row.get('Target Vsl')),
        'ism_points': n(row.get('ISM Points')), 'psc_det_history': n(row.get('PSC Det History')),
        'tonnage_client': s(row.get('Tonnage Client')),
        'vessel_type': s(row.get('Vessel Type')), 'age': n(row.get('Age')),
    })

print(f"{len(rows)} valid rows found. Clearing existing data...")
sb_request("DELETE", "inspection_history?id=neq.0")
print("Cleared. Inserting...")

BATCH, saved, skipped = 200, 0, 0
total = math.ceil(len(rows) / BATCH)
for i in range(0, len(rows), BATCH):
    batch = rows[i:i+BATCH]
    status, resp = sb_request("POST", "inspection_history", batch)
    if status in (200, 201):
        saved += len(batch)
    else:
        for row in batch:
            st, _ = sb_request("POST", "inspection_history", [row])
            if st in (200, 201): saved += 1
            else: skipped += 1
    batch_num = i // BATCH + 1
    if batch_num % 25 == 0 or batch_num == total:
        print(f"  {batch_num}/{total} batches — {saved} saved, {skipped} skipped")

print(f"\nDONE — {saved} rows inserted, {skipped} skipped")
